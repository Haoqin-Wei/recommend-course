"""
Convert the cleaned 3-sheet xlsx into 3 CSVs in data/uci/.

Usage:
    python scripts/import_term_data.py path/to/spring_sections_relational.xlsx

For multi-term support: run this once per fresh xlsx. It APPENDS to
existing CSVs, deduplicating by primary key. So a workflow like
    python scripts/import_term_data.py 2025_spring.xlsx
    python scripts/import_term_data.py 2025_fall.xlsx
    python scripts/import_term_data.py 2026_winter.xlsx
will produce three accumulated CSVs containing all terms.

Requires: openpyxl (pip install openpyxl)
"""

import sys
import csv
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Need openpyxl. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Map sheet → primary-key columns for deduplication on append.
SHEET_KEYS = {
    "sections":             ("section_id",),
    "section_instructors":  ("section_id", "instructor_name_raw"),
    "section_ge":           ("section_id", "ge_code"),
}


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not in {path}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows)]
    data = [list(r) for r in rows if any(c is not None for c in r)]
    return header, data


def write_csv_merged(
    out_path: Path,
    header: list[str],
    new_rows: list[list],
    key_cols: tuple[str, ...],
) -> tuple[int, int]:
    """Merge new_rows into out_path, deduplicating by key_cols.

    Returns (rows_added, rows_total).
    """
    existing: dict[tuple, list] = {}

    if out_path.exists():
        with out_path.open("r", encoding="utf-8") as f:
            reader = csv.reader(f)
            old_header = next(reader)
            if old_header != header:
                raise ValueError(
                    f"Header mismatch in {out_path}\n"
                    f"  old: {old_header}\n  new: {header}"
                )
            key_idx = [header.index(k) for k in key_cols]
            for row in reader:
                key = tuple(row[i] for i in key_idx)
                existing[key] = row

    key_idx = [header.index(k) for k in key_cols]
    added = 0
    for row in new_rows:
        row_str = ["" if v is None else str(v) for v in row]
        key = tuple(row_str[i] for i in key_idx)
        if key not in existing:
            added += 1
        existing[key] = row_str

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for row in existing.values():
            writer.writerow(row)

    return added, len(existing)


def main(xlsx_path: str, out_dir: str = "data/uci"):
    xlsx = Path(xlsx_path)
    out = Path(out_dir)
    if not xlsx.exists():
        print(f"Not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    for sheet, key_cols in SHEET_KEYS.items():
        header, data = read_sheet(xlsx, sheet)
        added, total = write_csv_merged(out / f"{sheet}.csv", header, data, key_cols)
        print(
            f"  {sheet}: read {len(data)} rows, "
            f"added {added} new → {out / f'{sheet}.csv'} (total {total})"
        )


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__, file=sys.stderr)
        sys.exit(1)
    main(*sys.argv[1:])
